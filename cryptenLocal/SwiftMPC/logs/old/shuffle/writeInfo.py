from curses.ascii import isdigit
import re
import openpyxl
from openpyxl import load_workbook

names = ['party0.txt', 'party1.txt', 'party2.txt']


#Reading all files
fA = open(names[0], 'r')
fB = open(names[1], 'r')
fC = open(names[2], 'r')

#Open required sheet in workbook
file_path = 'ShuffleBenchmarks.xlsx'
wb = load_workbook(file_path)

ws = wb["ShuffleSWIFT"]
#Get lines from all log files
lines1 = [line.rstrip() for line in fA]
lines2 = [line.rstrip() for line in fB]
lines3 = [line.rstrip() for line in fC]

col = "C"

curProc = 10
currentRow = 10
state = "P"

totalCommunication = 0
for i in range(0, len(lines1)):

    if "Preprocessing" in lines1[i]:
        state = "P"
    elif "Online" in lines1[i]:
        state = "O"

    if state == "P":
         ws["B" + str(currentRow)] = curProc
    
    if "My time" in lines1[i]:
        num1 = ""
        num2 = ""
        num3 = ""

        startIndex = len("My time: ")

        num1 = float(lines1[i][startIndex:])
        num2 = float(lines2[i][startIndex:])
        num3 = float(lines3[i][startIndex:])

        maxVal = max(num1, num2, num3)
        if state == "P":
            ws["C"+str(currentRow)] = maxVal
        elif state == "O":
            ws["E"+str(currentRow)] = maxVal
        
    if "Sent to" in lines1[i]:
        startIndex = len("Sent to 0: ")
        num1 = float(lines1[i][startIndex:])
        num2 = float(lines2[i][startIndex:])
        num3 = float(lines3[i][startIndex:])
        totalCommunication = totalCommunication + num1 + num2 + num3

        if "Sent to 2: " in lines1[i] and state == "P":
            ws["D"+str(currentRow)] = totalCommunication
            totalCommunication = 0
        elif "Sent to 2: " in lines1[i] and state == "O":
            ws["F"+str(currentRow)] = totalCommunication
            totalCommunication = 0
            currentRow = currentRow + 1
            curProc = curProc*10
     
wb.save(file_path)

    
